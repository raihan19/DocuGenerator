import os
import sys
import csv
import xlrd
from datetime import date
from openpyxl import load_workbook
import functions_weighting_modelling as f


if len(sys.argv) == 3:
    folderIn = str(sys.argv[1])
    folderOut = str(sys.argv[2])
else:
    folderIn = os.getcwd()
    folderOut = str("{}/documents/".format(os.getcwd()))

data = f.get_data_from_json(folderIn)
atlasList = f.atlas_list(data)
weaNo = []
weaName = []

for wtg in data['WEAs']:
    weaNo.append(wtg['fname'][3:4])
    weaName.append(wtg["Anlage"])

# Read from a file to get specific value from a cell B1
wb = load_workbook("{}/{}".format(folderIn, data['files']['Positionsdaten']))
sh = wb["Positionsdaten"]
name1 = sh["b1"].value

#opening the template and it's sheet
wb1 = load_workbook('{}/templates/Gewichtung_Modellierung.xlsx'.format(os.path.dirname(__file__)))
wb1_sh1 = wb1['LZ-Werte'] #opening first sheet
wb1_sh2 = wb1['Korrelationen'] #second sheet
wb1_sh3 = wb1['IW'] #third sheet
wb1_sh4 = wb1['Anz. Fehlwerte'] #fourth sheet
wb1_sh5 = wb1['Anz. Ertragswerte'] #fifth sheet
wb1_sh6 = wb1['Differenz'] #sixth sheet

countColumn = 3 #to track the column number

for refName in atlasList:
    i = 0
    j = 0
    for num in weaNo:
        #open the xls file
        wb2 = xlrd.open_workbook("{}/{}/wea{}.xls".format(folderIn, refName, num))

        '''LZ-Werte (long term values): WeaX.xls, Sheet 5 Cell H8 (Jahreserträge (Regression))'''
        wb2_sh1 = wb2.sheet_by_index(4) #fifth sheet
        wb2_sh1_val1 = wb2_sh1.cell_value(rowx=7, colx=7) #cell H8

        #write the refName
        row_val1 = wb1_sh1.cell(row=j+1, column=countColumn)
        row_val1.value = refName

        #write the values
        row_val2 = wb1_sh1.cell(row=i+3, column=countColumn)
        row_val2.value = wb2_sh1_val1
        row_val2.number_format = '0.000'

        #write the WEA name i.e WEA 1
        wea_name = wb1_sh1.cell(row=i+3, column=1)
        wea_name.value = weaName[i]

        '''Korrelationen (Correlations): WEAX.xls, Sheet 1, Cell H3 square them x^2'''
        wb2_sh2 = wb2.sheet_by_index(0) #first sheet
        wb2_sh2_val1 = wb2_sh2.cell_value(rowx=2, colx=7) #cell H3

        # write the refName
        row_val1 = wb1_sh2.cell(row=j + 1, column=countColumn)
        row_val1.value = refName

        # write the values
        row_val2 = wb1_sh2.cell(row=i + 3, column=countColumn)
        row_val2.value = wb2_sh2_val1**2
        row_val2.number_format = '0.000'

        # write the WEA name i.e WEA 1
        wea_name = wb1_sh2.cell(row=i + 3, column=1)
        wea_name.value = weaName[i]

        #opening ergebnis.xls file
        wb3 = xlrd.open_workbook("{}/{}/ergebnis.xls".format(folderIn, refName))

        '''IW (probability of error): ergebnis.xlsx, Sheet 1, B3(wea1), C3 (wea2)..'''
        wb3_sh1 = wb3.sheet_by_index(0) #first sheet
        wb3_sh1_val1 = wb3_sh1.cell_value(rowx=2, colx=i+1)

        # write the refName
        row_val1 = wb1_sh3.cell(row=j + 1, column=countColumn)
        row_val1.value = refName

        # write the values
        row_val2 = wb1_sh3.cell(row=i + 3, column=countColumn)
        row_val2.value = wb3_sh1_val1

        # write the WEA name i.e WEA 1
        wea_name = wb1_sh3.cell(row=i + 3, column=1)
        wea_name.value = weaName[i]

        '''This will loop through the csv file to check if the value is -9.0 and increment the counter'''
        countM = 0  # count missing
        countA = 0  # count all
        with open("{}/{}/{}".format(folderIn, refName, data['WEAs'][i]['fname']), newline='') as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                countA += 1
                if row["WEA {}".format(weaNo[i])] == '-9.0':
                    countM += 1
        countNM = countA - countM  # count non-missing

        '''Anz. Fehlwerte (Number of missing values): WEAX.xls, “-9” in weaX.csv'''
        #write the refName
        row_val1 = wb1_sh4.cell(row=j+1, column=countColumn)
        row_val1.value = refName

        #write the values
        row_val2 = wb1_sh4.cell(row=i + 3, column=countColumn)
        row_val2.value = countM
        #print(countM)

        row_val3 = wb1_sh4.cell(row=1, column=1)
        row_val3.value = 'Fehlwerte in den Ertragsdaten'

        #write WEA name
        wea_name = wb1_sh4.cell(row=i + 3, column=1)
        wea_name.value = weaName[i]

        '''Anz. Ertragswerte (Number of energy yield values): WEAX.xls, not “-9” in weaX.csv'''
        # write the refName
        row_val1 = wb1_sh5.cell(row=j + 1, column=countColumn)
        row_val1.value = refName

        # write the values
        row_val2 = wb1_sh5.cell(row=i + 3, column=countColumn)
        row_val2.value = countNM

        row_val3 = wb1_sh5.cell(row=1, column=1)
        row_val3.value = 'Verwendete Ertragswerte'

        # write WEA name
        wea_name = wb1_sh5.cell(row=i + 3, column=1)
        wea_name.value = weaName[i]

        '''Differenz (difference): ergebnis.xlsx, Sheet 1, B6(wea1), C6 (wea2)…'''
        wb3_sh1 = wb3.sheet_by_index(0)  # first sheet
        wb3_sh1_val1 = wb3_sh1.cell_value(rowx=5, colx=i + 1)

        # write the refName
        row_val1 = wb1_sh6.cell(row=j + 1, column=countColumn)
        row_val1.value = refName

        # write the values
        row_val2 = wb1_sh6.cell(row=i + 3, column=countColumn)
        row_val2.value = wb3_sh1_val1

        # write the WEA name i.e WEA 1
        wea_name = wb1_sh6.cell(row=i + 3, column=1)
        wea_name.value = weaName[i]

        i += 1
    countColumn += 1

try:
    if not os.path.isdir(folderOut):
        os.makedirs("{}".format(folderOut))
except OSError as e:
    print(e)

wb1.save("{}/{}_{}_Gewichtung_Modellierung.xlsx".format(folderOut, date.today(), name1))



