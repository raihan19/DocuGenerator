import os
import sys
import xlrd
import xlsxwriter
from datetime import date
import openpyxl
from openpyxl import load_workbook
import functions_compare_index as f
import pandas as pd

if len(sys.argv) == 3:
    folderIn = str(sys.argv[1])
    folderOut = str(sys.argv[2])
else:
    folderIn = os.getcwd()
    folderOut = str("{}/documents/".format(os.getcwd()))

#WTG number
weaNo = 1

data = f.get_data_from_json(folderIn)
atlasList = f.atlas_list(data)

# Read from a file to get specific value from a cell B1
wb = load_workbook("{}/{}".format(folderIn, data['files']['Positionsdaten']))
sh = wb["Positionsdaten"]
name1 = sh["b1"].value

#formatting the name of the xlsx file
wBook1 = xlsxwriter.Workbook("{}/{}_{}_Vergleich_Indizes_Trend.xlsx".format(folderOut, date.today(), name1)) #date.today() will give the updated date when the script will be executed

#name the 1st sheet 'Daten' and 2nd sheet 'Diagramme' and save the location of the object
wBook1_sheet1 = wBook1.add_worksheet("Daten")
wBook1_sheet2 = wBook1.add_worksheet("Diagramme")

# format the color of the background
multiple_cell_color = wBook1.add_format({'bg_color': '#DDDDDD'})

#color multiple cells
for i in range(3):
    for j in range(100):
        wBook1_sheet1.write(i, j, '', multiple_cell_color)

#countColumn helps to track column number
countColumn = 0

# write on specific cell
wBook1_sheet1.write(1, 0, "Windindex", multiple_cell_color) #A2
wBook1_sheet1.write(2, 0, "Zeitraum", multiple_cell_color)  #A3
countColumn += 1

i = 0
j = 1

countMax = 0 #count the maximum input value from a particular atlas

'''For every atlas in first column put the dates and in subsequent column put the value. 
i, j tracks down the column, k tracks the row. countMax is created to tract maximum number of entry. 
count tracks entries for every atlas'''

all_refs = []

for refName in atlasList:
    refData = {}
    refData["name"] = refName

    count = 0 #number of values in the column
    wBook1_sheet1.write(2, i+1, f.atlasName(data)[refName], multiple_cell_color)
    i += 1
    countColumn += 1

    #open and read the data from the source file
    wb = xlrd.open_workbook("{}/{}/wea{}.xls".format(folderIn, refName, weaNo))
    sheet = wb.sheet_by_index(2)        #accessed in the source sheet

    date_value = {}

    k = 0
    #read from 2nd column, 2nd row
    for val in sheet.col_values(1, 1):
        #read from input file
        date_value[sheet.cell_value(rowx=k+1, colx=0)] = val
        count += 1
        k += 1

    refData["data"] = date_value
    all_refs.append(refData)

first_date = f.get_first_common_date(all_refs)

for refName in atlasList:
    #print(refName)
    refData = f.get_ref_data_first_date(all_refs, refName, first_date)
    count = len(refData)
    if count > countMax:
        countMax = count
        new_date_value = refData
    date_value = refData

    #formatting date and value
    date_format = wBook1.add_format({'num_format': 'mm/dd/yyyy'}) #01/02/1999 for the date format
    val_format = wBook1.add_format({'num_format': '0.00'}) #for the value format

    #write the date and value on the generated file
    k = 0
    for dateV in sorted(date_value.keys()):
        #write to output file
        wBook1_sheet1.write(k+3, 0, dateV, date_format) # 01/01/2000, fourth row, first column
        wBook1_sheet1.write(k+3, j, date_value[dateV], val_format)
        k += 1
    j += 1

wBook1_sheet1.write(2, countColumn, 'Abweichung', multiple_cell_color)
countColumn += 1

#create the excel file
try:
    if not os.path.isdir(folderOut):
        os.makedirs("{}".format(folderOut))
except OSError as e:
    print(e)

wBook1.close()

#load the workbook
wb_main = load_workbook("{}/{}_{}_Vergleich_Indizes_Trend.xlsx".format(folderOut, date.today(), name1))
sh = wb_main["Daten"]
# create list of specific alphabets, 98 = b, 99 = c...
specific_alphabet1 = list(map(chr, range(98, 98+len(f.atlas_list(data)))))

val_row_list = []

'''create a list of list where each list contains entry from each row'''
for i in range(countMax):
    val_row = []
    for j in range(len(f.atlas_list(data))):
        if sh["{}{}".format(specific_alphabet1[j], i+4)].value == None:
            continue
        val_row.append(sh["{}{}".format(specific_alphabet1[j], i+4)].value)
    val_row_list.append(val_row)

'''Get the deviation from each row'''
diff_list1 = f.deviation_list(countMax, val_row_list)

# Get maximum deviation
maxDiff = max(diff_list1)

color_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor='FFFF00')
#write the deviation in the file
for i in range(len(diff_list1)):
    row_val1 = sh.cell(row=i+4, column=countColumn)
    row_val1.value = diff_list1[i]
    if diff_list1[i] > 30:
        row_val1.fill = color_fill
    row_val1.number_format = '0.00'

countColumn += 1
row_val1 = sh.cell(3, column=countColumn)
row_val1.value = 'max Abweichung'
row_val1 = sh.cell(4, column=countColumn)
row_val1.value = maxDiff

'''To calculate the average of 12 months'''
avg_row_list = []
for i in range(len(f.atlas_list(data))):
    avg_row = []
    for j in range(countMax):
        if sh["{}{}".format(specific_alphabet1[i], j+4)].value == None:
            continue
        avg_row.append(sh["{}{}".format(specific_alphabet1[i], j+4)].value)
    avg_row_list.append(avg_row)

countColumn += 2
tempCounterColumn = countColumn
columnPlot1 = countColumn
row_val1 = sh.cell(2, column = countColumn)
row_val1.value = "gleitendes 12-Monatsmittel"
row_val1 = sh.cell(3, column = countColumn)
row_val1.value = "Zeitraum"

countColumn += 1

#inserting the atlas name for gleitendes 12-Monatsmittel
i = 0
for refName in atlasList:
    row_val1 = sh.cell(3, column=countColumn)
    row_val1.value = f.atlasName(data)[refName]
    row_val2 = sh.cell(14, column=countColumn)
    row_val2.value = f.atlasName(data)[refName]
    countColumn += 1

    j = 0
    for dateV in sorted(new_date_value.keys()):
        date_val1 = sh.cell(4+j, column=tempCounterColumn)
        date_val1.value = dateV
        date_val1.number_format = "mm/dd/yyyy"
        date_val2 = sh.cell(15 + j, column=countColumn - 1)
        if j > len(avg_row_list[i]) - 12:
            j += 1
            continue
        date_val2.value = (avg_row_list[i][j] + avg_row_list[i][j+1] + avg_row_list[i][j+2] + avg_row_list[i][j+3] + avg_row_list[i][j+4] + avg_row_list[i][j+5] + avg_row_list[i][j+6] + avg_row_list[i][j+7] + avg_row_list[i][j+8] + avg_row_list[i][j+9] + avg_row_list [i][j+10] + avg_row_list[i][j+11]) / 12.0
        date_val2.number_format = "0.00"
        j += 1
    i += 1

val_row_list = []
tempCounterColumn = countColumn - len(f.atlas_list(data))
for i in range(countMax - 11):
    val_row = []
    countColumn = tempCounterColumn
    for j in range(len(f.atlas_list(data))):
        row_val3 = sh.cell(15 + i, column=countColumn)
        if row_val3.value == None:
            countColumn += 1
            continue
        val_row.append(row_val3.value)
        countColumn += 1
    val_row_list.append(val_row)

diff_list1 = f.deviation_list(countMax-11, val_row_list)
maxDiff = max(diff_list1)

for i in range(len(diff_list1)):
    row_val1 = sh.cell(row=i+15, column=countColumn)
    row_val1.value = diff_list1[i]
    if diff_list1[i] > 7:
        row_val1.fill = color_fill
    row_val1.number_format = "0.00"

row_val1 = sh.cell(3, column=countColumn)
row_val1.value = "Abweichung"

for i in range(11):
    row_val1 = sh.cell(i+4, column=countColumn)
    row_val1.value = 0
    row_val1.number_format = "0.00"

countColumn += 1
row_val1 = sh.cell(3, column=countColumn)
row_val1.value = 'max Abweichung'
row_val1 = sh.cell(4, column=countColumn)
row_val1.value = maxDiff
#row_val1.number_format = '0.00'

'''To calculate the average of 24 months'''

countColumn += 2
tempCounterColumn = countColumn
columnPlot2 = countColumn
row_val1 = sh.cell(2, column = countColumn)
row_val1.value = "gleitendes 24-Monatsmittel"
row_val1 = sh.cell(3, column = countColumn)
row_val1.value = "Zeitraum"

countColumn += 1

#inserting the atlas name for gleitendes 24-Monatsmittel
i = 0
for refName in atlasList:
    row_val1 = sh.cell(3, column=countColumn)
    row_val1.value = f.atlasName(data)[refName]
    row_val2 = sh.cell(26, column=countColumn)
    row_val2.value = f.atlasName(data)[refName]
    countColumn += 1

    j = 0
    for dateV in sorted(new_date_value.keys()):
        date_val1 = sh.cell(4+j, column=tempCounterColumn)
        date_val1.value = dateV
        date_val1.number_format = "mm/dd/yyyy"
        date_val2 = sh.cell(27 + j, column=countColumn - 1)
        if j > len(avg_row_list[i]) - 24:
            j += 1
            continue
        date_val2.value = (avg_row_list[i][j] + avg_row_list[i][j+1] + avg_row_list[i][j+2] + avg_row_list[i][j+3] + avg_row_list[i][j+4] + avg_row_list[i][j+5] + avg_row_list[i][j+6] + avg_row_list[i][j+7] + avg_row_list[i][j+8] + avg_row_list[i][j+9] + avg_row_list [i][j+10] + avg_row_list[i][j+11] + avg_row_list[i][j+12] + avg_row_list[i][j+13] + avg_row_list[i][j+14] + avg_row_list[i][j+15] + avg_row_list[i][j+16] + avg_row_list[i][j+17] + avg_row_list[i][j+18] + avg_row_list[i][j+19] + avg_row_list[i][j+20] + avg_row_list[i][j+21] + avg_row_list[i][j+22] + avg_row_list[i][j+23]) / 24.0
        date_val2.number_format = "0.00"
        j += 1
    i += 1

val_row_list = []
tempCounterColumn = countColumn - len(f.atlas_list(data))
for i in range(countMax - 23):
    val_row = []
    countColumn = tempCounterColumn
    for j in range(len(f.atlas_list(data))):
        row_val3 = sh.cell(27 + i, column=countColumn)
        if row_val3.value == None:
            countColumn += 1
            continue
        val_row.append(row_val3.value)
        countColumn += 1
    val_row_list.append(val_row)

diff_list1 = f.deviation_list(countMax-23, val_row_list)
maxDiff = max(diff_list1)

for i in range(len(diff_list1)):
    row_val1 = sh.cell(row=i+27, column=countColumn)
    row_val1.value = diff_list1[i]
    if diff_list1[i] > 4:
        row_val1.fill = color_fill
    row_val1.number_format = "0.00"

row_val1 = sh.cell(3, column=countColumn)
row_val1.value = "Abweichung"

for i in range(23):
    row_val1 = sh.cell(i+4, column=countColumn)
    row_val1.value = 0
    row_val1.number_format = "0.00"

countColumn += 1
row_val1 = sh.cell(3, column=countColumn)
row_val1.value = 'max Abweichung'
row_val1 = sh.cell(4, column=countColumn)
row_val1.value = maxDiff

'''Calculate Yearly Average'''

'''For every atlas in first column put the dates and in subsequent column put the value. 
i, j tracks down the column, k tracks the row. countMax is created to tract maximum number of entry. 
count tracks entries for every atlas'''

all_refs = []

for refName in atlasList:
    refData = {}
    refData["name"] = refName
    i += 1

    #open and read the data from the source file
    wb_1 = xlrd.open_workbook("{}/{}/wea{}.xls".format(folderIn, refName, weaNo))
    sheet = wb_1.sheet_by_index(2)        #accessed in the source sheet

    date_value = {}

    k = 0
    #read from 2nd column, 2nd row
    for val in sheet.col_values(17, 2):
        #read from input file
        if sheet.cell_value(rowx=k+2, colx=4) != '':
            keyV = int(sheet.cell_value(rowx=k+2, colx=4))
        else:
            continue
        date_value[keyV] = val
        k += 1

    refData["data"] = date_value
    all_refs.append(refData)

first_date = f.get_first_common_date(all_refs)
date_value_list = []
for refName in atlasList:
    refData = f.get_ref_data_first_date(all_refs, refName, first_date)
    date_value = refData
    date_value_list.append(date_value)

countColumn += 2
tempCounterColumn = countColumn
columnPlot3 = countColumn
row_val1 = sh.cell(2, column = countColumn)
row_val1.value = 'Jahresmittel'
row_val1 = sh.cell(3, column = countColumn)
row_val1.value = 'Jahr'

'''Remove year whose average doesn't include 12 months'''
df = pd.read_excel("{}/{}_{}_Vergleich_Indizes_Trend.xlsx".format(folderOut, date.today(), name1))
date_val = df['Unnamed: 0'].to_dict()
countCheck = 1
for refName in atlasList:
    val_pair_list = {}
    val_pair = df['Unnamed: {}'.format(countCheck)].to_dict()
    j = 0
    for i in range(len(val_pair)):
        if pd.isnull(val_pair[i]):
            continue
        else:
            val_pair_list[j] = val_pair[i]
            j += 1

    date_check_val = []
    i = 1
    for key in date_val:
        if key < 2:
            continue
        if i == len(val_pair_list):
            break
        date_check_val.append(date_val[key].strftime('%m/%Y'))
        i += 1

    if int(date_check_val[-1][0:2]) < 13:
        rmvYear = int(date_check_val[-1][3:])

    for dateV in sorted(date_value_list[countCheck-1].keys()):
        date_value_list[countCheck-1].pop(rmvYear, None)
    countCheck += 1

countColumn += 1
m = 0
newCountMax = 0
for refName in atlasList:
    newCount = 0
    row_val1 = sh.cell(3, column = countColumn)
    row_val1.value = f.atlasName(data)[refName]
    countColumn += 1

    j = 0
    for dateV in sorted(date_value_list[m].keys()):
        newCount += 1
        date_val1 = sh.cell(4 + j, column=tempCounterColumn)
        date_val1.value = dateV
        date_val2 = sh.cell(4 + j, column=countColumn - 1)
        date_val2.value = date_value_list[m][dateV]
        date_val2.number_format = '0.00'
        j += 1
    if newCountMax < newCount:
        newCountMax = newCount
    m += 1

# tempCounterColumn = countColumn - len(f.atlas_list(data))

wb_main.save("{}/{}_{}_Vergleich_Indizes_Trend.xlsx".format(folderOut, date.today(), name1))


'''Line chart'''
'''Plot 1'''
val_row_list = []
'''create a list of list where each list contains entry from each row'''
for i in range(len(f.atlas_list(data))):
    val_row = []
    for j in range(countMax):
        if sh["{}{}".format(specific_alphabet1[i], j+4)].value == None:
            continue
        val_row.append(sh["{}{}".format(specific_alphabet1[i], j+4)].value)
    val_row_list.append(val_row)

df = pd.read_excel("{}/{}_{}_Vergleich_Indizes_Trend.xlsx".format(folderOut, date.today(), name1))
date_val = df['Unnamed: 0'].to_dict()

title_name = 'Monatsverlauf Ertragsindex'
limit = 2
figNo = 1
f.figure_1(val_row_list, atlasList, date_val, title_name, limit, figNo)

'''Plot 2'''
date_val = df['Unnamed: {}'.format(columnPlot1-1)].to_dict()
val_row_list = f.column_wise_value(data, countMax-11, columnPlot1, sh, 15)

title_name = 'gleitendes 12-Monatsmittel'
limit = 13
figNo = 2
f.figure_1(val_row_list, atlasList, date_val, title_name, limit, figNo)

'''Plot 3'''
date_val = df['Unnamed: {}'.format(columnPlot2-1)].to_dict()
val_row_list = f.column_wise_value(data, countMax-23, columnPlot2, sh, 27)

title_name = 'gleitendes 24-Monatsmittel'
limit = 25
figNo = 3
f.figure_1(val_row_list, atlasList, date_val, title_name, limit, figNo)

'''Plot 4'''
date_val = df['Unnamed: {}'.format(columnPlot3-1)].to_dict()
new_date_val = {}
track = 0
for key in date_val:
    if track == newCountMax+2:
        break
    new_date_val[key] = date_val[key]
    track += 1

val_row_list = f.column_wise_value(data, newCountMax, columnPlot3, sh, 4)
f.figure_2(atlasList, new_date_val, val_row_list)
